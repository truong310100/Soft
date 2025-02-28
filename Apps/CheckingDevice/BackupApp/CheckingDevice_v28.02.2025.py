import os
import re
import sys
import time
import json
import ctypes
import requests
import datetime
import subprocess
import webbrowser
import tkinter as tk
from tkinter import ttk
from bs4 import BeautifulSoup
from tkinter import messagebox
from openpyxl import load_workbook

UPLOAD_URL = "https://chuky.hunghau.org/apis_v1/bienban/upload"
unikey_path = r"BackupApp\\UniKeyNT.exe"  # Đường dẫn tới UniKey
cmd_check_health_status_disk = r'.\\CrytalDiskInfo\DiskInfo64.exe /CopyExit'
json_path_json = r'BackupApp\\list_data_localhost.json'
read_log_disk_info = 'CrytalDiskInfo/DiskInfo.txt'
cmd_create_report_pin = 'powercfg /batteryreport'
cmd_start_crytaldiskinfo = 'start CrytalDiskInfo\\DiskInfo64.exe'
cmd_read_report_pin = 'battery-report.html'
cmd_rename = 'start BackupApp\\Rename_PC.cmd'
web_check_screen = 'https://www.youtube.com/watch?v=d9LdhipeZ40'
web_check_speaker='https://mymictest.com/vi/speaker-test/'
web_check_keyboard='https://en.key-test.ru/'
web_check_camera='https://www.onlinemictest.com/webcam-test/'
web_check_micro='https://www.onlinemictest.com/'
# web_sharepoint = 'https://hunghaugroup.sharepoint.com/:f:/g/HHH.HungHauHoldings/Es1sMfxQCtJMvZjz8MH6trMB9do2xRhMbPQbPJSXy1pjww?e=BReDQB'
web_sharepoint = 'https://chuky.hunghau.org/bienban'
write_data_to_excel = 'BackupApp/BanGiao.xlsx'

def run_as_admin():
    if not ctypes.windll.shell32.IsUserAnAdmin():
        ctypes.windll.shell32.ShellExecuteW(None, "runas", sys.executable, " ".join(sys.argv), None, 0)
        sys.exit()
run_as_admin()

def run_unikey_as_admin():
    try:
        subprocess.run(["powershell", "Start-Process", unikey_path, "-Verb", "RunAs"], shell=True)
    except Exception as e:
        print(f"Lỗi khi chạy UniKey: {e}")
run_unikey_as_admin()

def load_support_data():
    try:
        with open(json_path_json, "r", encoding="utf-8") as file:
            data = json.load(file)
            return {
                "names": [person["name"] for person in data], 
                "working_locations": [person["working_location"] for person in data],
                "emails": [person["email"] for person in data] 
            }
    except Exception as e:
        print(f"Lỗi khi đọc file JSON: {e}")
        return {"names": [], "working_locations": [], "emails": []}
support_data = load_support_data()
support_names = support_data["names"]
support_emails = support_data["emails"]
support_working_locations = support_data["working_locations"]

def check_pin():
    os.system(cmd_read_report_pin)
    update_battery_status()

def update_battery_status():
    try:
        with open(cmd_read_report_pin, 'r', encoding='utf-8') as f:
            soup = BeautifulSoup(f.read(), 'html.parser')

        design_capacity = int(soup.find('td', string='DESIGN CAPACITY').find_next_sibling('td').text.strip().replace(',', '').replace(' mWh', ''))
        full_charge_capacity = int(soup.find('td', string='FULL CHARGE CAPACITY').find_next_sibling('td').text.strip().replace(',', '').replace(' mWh', ''))
        remaining_capacity = full_charge_capacity / design_capacity * 100
        entry_pin.delete(0, tk.END)
        entry_pin.insert(0, f"{remaining_capacity:.2f}%")
    except Exception as e:
        messagebox.showerror("Lỗi", f"Không thể đọc báo cáo pin: {e}")
    
def check_driver():
    urls = [
        web_check_screen,
        web_check_speaker,
        web_check_keyboard,
        # web_check_camera,
        web_check_micro
    ]
    os.system('start microsoft.windows.camera:')
    for url in urls:
        webbrowser.open(url)
        time.sleep(1)

def check_harddisk():
    os.system(cmd_check_health_status_disk)
    os.system(cmd_start_crytaldiskinfo)
    update_disk_health()
    
def read_disk_health():
    os.system(cmd_check_health_status_disk)
    if not os.path.exists(read_log_disk_info):
        os.system(cmd_start_crytaldiskinfo)
    with open(read_log_disk_info, "r", encoding="ascii") as file:
        lines = file.readlines()
    for line in lines:
        if "Health Status" in line:
            return line.strip() 
    return "Không xác định"

def update_disk_health():
    disk_health = read_disk_health() or "Không xác định"
    entry_note3.delete(0, tk.END)
    entry_note3.insert(0, disk_health)
    
def submit_form():
    input_user_Support = support_combobox.get()
    input_user_name = entry_name.get()
    input_department = entry_department.get()
    text_note_1 = entry_note1.get()
    text_note_2 = entry_note2.get()
    text_note_3 = entry_note3.get()
    
    type_bb = type_var.get()
    selected_support = support_combobox.get()
    if selected_support in support_names:
        index = support_names.index(selected_support)
        working_location = support_working_locations[index]
    else:
        working_location = "Không xác định"
        
    # Lấy thông tin pin từ entry_pin
    battery_status = entry_pin.get()
    
    # Lấy thông tin hệ thống
    hostname = entry_hostname.get().strip()
    system_model = subprocess.run(['wmic', 'csproduct', 'get', 'name'], capture_output=True, text=True).stdout.strip().split('\n')[2]
    serial = subprocess.run(['wmic', 'bios', 'get', 'serialnumber'], capture_output=True, text=True).stdout.strip().split('\n')[2]
    CPU_Name = subprocess.run(['wmic', 'cpu', 'get', 'name'], capture_output=True, text=True).stdout.strip().split('\n')[2]
    
    #RAM
    capacities = subprocess.run(['wmic', 'MemoryChip', 'get', 'Capacity'], capture_output=True, text=True).stdout.strip().split('\n')[2:]
    speeds = subprocess.run(['wmic', 'MemoryChip', 'get', 'speed'], capture_output=True, text=True).stdout.strip().split('\n')[2:]
    manufacturers = subprocess.run(['wmic', 'MemoryChip', 'get', 'manufacturer'], capture_output=True, text=True).stdout.strip().split('\n')[2:]

    ram_info = [(cap.strip(), speed.strip(), man.strip()) for cap, speed, man in zip(capacities, speeds, manufacturers) if cap]
    rams = []
    for i, (capacity, speed, manufacturer) in enumerate(ram_info):
        Ram_OS = (f"RAM {i + 1}: {capacity:.1}GB {speed} {manufacturer}")
        rams.append(Ram_OS)
        print(Ram_OS)
    Ram_OS = '\n'.join(rams)
    
    #Disk
    cmd = 'Get-PhysicalDisk | Select-Object MediaType, Manufacturer, Model, Size'
    result = subprocess.run(['powershell', '-Command', cmd], capture_output=True, text=True)

    if result.stderr:
        print(f"Error: {result.stderr.strip()}")
    else:
        output = result.stdout.strip().splitlines()[2:]
        print('Disk:')
        disks = []
        i = 0
        for line in output:
            i += 1 
            parts = line.strip().split()
            media_type = parts[0]
            manufacturer = parts[1]
            model = ' '.join(parts[2:-1])
            size_bytes = int(parts[-1])
            size_gb = size_bytes / 2**30
            Disk = (f"Media {i}: {media_type}, Manufacturer: {manufacturer}, Model: {model}, Size: {size_gb:.2f} GB")
            disks.append(Disk)
            print(Disk)
        Disk = '\n'.join(disks)
    
    hard_ware = f"{CPU_Name}\n{Ram_OS}\n{Disk}"
    # Ghi vào file Excel
    workbook = load_workbook(write_data_to_excel)
    sheet = workbook['BienBan']
    sheet['B10'] = input_user_Support
    sheet['D10'] = working_location
    sheet['D12'] = input_department
    sheet['A10'], sheet['B12'], sheet['D14'] = ('Người giao:' if type_bb == 'Giao' else 'Người nhận:'), input_user_name, hostname
    sheet['E17'], sheet['B17'], sheet['C17'] = serial, system_model, hard_ware
    sheet['F17'] = f"Tình trạng thân máy: {text_note_1}\nKết nối khác: {text_note_2}\nTuổi thọ ổ cứng: {text_note_3}\nPin còn lại: {battery_status}"

    # Lưu và mở file
    if not os.path.exists('Export'):
        os.makedirs('Export')
    filename = f'Export/BB-{type_bb}-{input_user_name}-{datetime.date.today().strftime("%d%m%Y")}.xlsx'
    workbook.save(filename)
    os.system(f'start "" "{filename}"')
    os.system(f'start "" "{web_sharepoint}"')
    # os.remove(cmd_read_report_pin)
    send_file()
    
def fast_send_file():
    input_user_Support = support_combobox.get()
    input_user_name = entry_name.get()
    input_department = entry_department.get()
    text_note_1 = entry_note1.get()
    text_note_2 = entry_note2.get()
    text_note_3 = entry_note3.get()
    
    type_bb = type_var.get()
    selected_support = support_combobox.get()
    if selected_support in support_names:
        index = support_names.index(selected_support)
        working_location = support_working_locations[index]
    else:
        working_location = "Không xác định"
        
    # Lấy thông tin pin từ entry_pin
    battery_status = entry_pin.get()
    
    # Lấy thông tin hệ thống
    hostname = entry_hostname.get().strip()
    system_model = subprocess.run(['wmic', 'csproduct', 'get', 'name'], capture_output=True, text=True).stdout.strip().split('\n')[2]
    serial = subprocess.run(['wmic', 'bios', 'get', 'serialnumber'], capture_output=True, text=True).stdout.strip().split('\n')[2]
    CPU_Name = subprocess.run(['wmic', 'cpu', 'get', 'name'], capture_output=True, text=True).stdout.strip().split('\n')[2]
    
    #RAM
    capacities = subprocess.run(['wmic', 'MemoryChip', 'get', 'Capacity'], capture_output=True, text=True).stdout.strip().split('\n')[2:]
    speeds = subprocess.run(['wmic', 'MemoryChip', 'get', 'speed'], capture_output=True, text=True).stdout.strip().split('\n')[2:]
    manufacturers = subprocess.run(['wmic', 'MemoryChip', 'get', 'manufacturer'], capture_output=True, text=True).stdout.strip().split('\n')[2:]

    ram_info = [(cap.strip(), speed.strip(), man.strip()) for cap, speed, man in zip(capacities, speeds, manufacturers) if cap]
    rams = []
    for i, (capacity, speed, manufacturer) in enumerate(ram_info):
        Ram_OS = (f"RAM {i + 1}: {capacity:.1}GB {speed} {manufacturer}")
        rams.append(Ram_OS)
        print(Ram_OS)
    Ram_OS = '\n'.join(rams)
    
    #Disk
    cmd = 'Get-PhysicalDisk | Select-Object MediaType, Manufacturer, Model, Size'
    result = subprocess.run(['powershell', '-Command', cmd], capture_output=True, text=True)

    if result.stderr:
        print(f"Error: {result.stderr.strip()}")
    else:
        output = result.stdout.strip().splitlines()[2:]
        print('Disk:')
        disks = []
        i = 0
        for line in output:
            i += 1 
            parts = line.strip().split()
            media_type = parts[0]
            manufacturer = parts[1]
            model = ' '.join(parts[2:-1])
            size_bytes = int(parts[-1])
            size_gb = size_bytes / 2**30
            Disk = (f"Media {i}: {media_type}, Manufacturer: {manufacturer}, Model: {model}, Size: {size_gb:.2f} GB")
            disks.append(Disk)
            print(Disk)
        Disk = '\n'.join(disks)
    
    hard_ware = f"{CPU_Name}\n{Ram_OS}\n{Disk}"
    # Ghi vào file Excel
    workbook = load_workbook(write_data_to_excel)
    sheet = workbook['BienBan']
    sheet['B10'] = input_user_Support
    sheet['D10'] = working_location
    sheet['D12'] = input_department
    sheet['A10'], sheet['B12'], sheet['D14'] = ('Người giao:' if type_bb == 'Giao' else 'Người nhận:'), input_user_name, hostname
    sheet['E17'], sheet['B17'], sheet['C17'] = serial, system_model, hard_ware
    sheet['F17'] = f"Tình trạng thân máy: {text_note_1}\nKết nối khác: {text_note_2}\nTuổi thọ ổ cứng: {text_note_3}\nPin còn lại: {battery_status}"

    # Lưu và mở file
    if not os.path.exists('Export'):
        os.makedirs('Export')
    filename = f'Export/BB-{type_bb}-{input_user_name}-{datetime.date.today().strftime("%d%m%Y")}.xlsx'
    workbook.save(filename)
    # os.system(f'start "" "{filename}"')
    os.system(f'start "" "{web_sharepoint}"')
    # os.remove(cmd_read_report_pin)
    send_file()

def send_file():
    selected_support = support_combobox.get()
    if selected_support in support_names:
        index = support_names.index(selected_support)
        upload_by_email = support_emails[index]
    else:
        upload_by_email = selected_support

    filename = f'Export/BB-{type_var.get()}-{entry_name.get()}-{datetime.date.today().strftime("%d%m%Y")}.xlsx'
    if not os.path.exists(filename):
        messagebox.showerror("Lỗi", "File không tồn tại.")
        return

    files = {'file': open(filename, 'rb')}
    data = {'UploadBy': upload_by_email}

    try:
        response = requests.post(UPLOAD_URL, files=files, data=data)
        response.raise_for_status()
        messagebox.showinfo("Success", "Gửi file thành công!")
        os.remove(cmd_read_report_pin)
    except requests.exceptions.RequestException as e:
        messagebox.showerror("Error", f"Error uploading file: {e}")

def rename_pc():
    new_hostname = entry_hostname.get().strip()
    if not new_hostname:
        return messagebox.showerror("Lỗi", "Vui lòng nhập tên máy mới.")
    if messagebox.askyesno("Xác nhận", f"Đổi tên máy thành '{new_hostname}'?"):
        subprocess.run(f'wmic computersystem where name="%COMPUTERNAME%" call rename name="{new_hostname}"', shell=True)
        messagebox.showinfo("Thành công", "Tên máy đã được đổi, hệ thống sẽ khởi động lại.")
        os.system("shutdown /r /t 5")
        
# Giao diện tkinter
root = tk.Tk()
icon = tk.PhotoImage(file='BackupApp/logo-HHH.png')
root.tk.call('wm', 'iconphoto', root._w, icon)
root.title("Checking Device")
root.configure(bg="#f0f0f0")

# Tạo style cho các widget
style = ttk.Style()
style.configure("TLabel", background="#f0f0f0", font=("Arial", 10))
style.configure("TButton", font=("Arial", 10), padding=5)
style.configure("TEntry", font=("Arial", 10), padding=5)
style.configure("TCombobox", font=("Arial", 10), padding=5)

button_width = 20  
ttk.Button(root, text="Kiểm tra kết nối", command=check_driver, style="TButton").grid(row=5, column=2, padx=10, pady=5, sticky="ew")
ttk.Button(root, text="Đổi tên máy", command=rename_pc, style="TButton").grid(row=6, column=2, padx=10, pady=5, sticky="ew")
ttk.Button(root, text="Kiểm tra ổ cứng", command=check_harddisk, style="TButton").grid(row=7, column=2, padx=10, pady=5, sticky="ew")
ttk.Button(root, text="Kiểm tra pin", command=check_pin, style="TButton").grid(row=8, column=2, padx=10, pady=5, sticky="ew")
root.columnconfigure(2, weight=1)


# Họ & tên người hỗ trợ (sử dụng Combobox)
ttk.Label(root, text="Họ & tên người hỗ trợ:", style="TLabel").grid(row=0, column=0, padx=10, pady=5, sticky="w")
support_var = tk.StringVar()
support_combobox = ttk.Combobox(root, textvariable=support_var, values=support_names, style="TCombobox")
support_combobox.grid(row=0, column=1, padx=10, pady=5, sticky="ew", ipadx=100)
support_combobox.current(0)

# Loại biên bản
ttk.Label(root, text="Loại biên bản:", style="TLabel").grid(row=1, column=0, padx=10, pady=5, sticky="w")
type_var = tk.StringVar()
type_select = ttk.Combobox(root, textvariable=type_var, values=["Giao", "Nhận"], style="TCombobox")
type_select.grid(row=1, column=1, padx=10, pady=5, sticky="ew", ipadx=100)
type_select.current(0)

# Họ & tên người dùng
ttk.Label(root, text="Họ & tên người dùng:", style="TLabel").grid(row=2, column=0, padx=10, pady=5, sticky="w")
entry_name = ttk.Entry(root, style="TEntry")
entry_name.grid(row=2, column=1, padx=10, pady=5, sticky="ew", ipadx=100)

# Chức vụ - Phòng ban người dùng
ttk.Label(root, text="Chức vụ - Phòng ban người dùng:", style="TLabel").grid(row=3, column=0, padx=10, pady=5, sticky="w")
entry_department = ttk.Entry(root, style="TEntry")
entry_department.grid(row=3, column=1, padx=10, pady=5, sticky="ew", ipadx=100)

# Tình trạng thân máy
ttk.Label(root, text="Tình trạng thân máy:", style="TLabel").grid(row=4, column=0, padx=10, pady=5, sticky="w")
entry_note1 = ttk.Entry(root, style="TEntry")
entry_note1.grid(row=4, column=1, padx=10, pady=5, sticky="ew", ipadx=100)

# Kết nối khác
ttk.Label(root, text="Kết nối khác:", style="TLabel").grid(row=5, column=0, padx=10, pady=5, sticky="w")
entry_note2 = ttk.Entry(root, style="TEntry")
entry_note2.grid(row=5, column=1, padx=10, pady=5, sticky="ew", ipadx=100)

# Tên máy người dùng
ttk.Label(root, text="Tên máy người dùng:", style="TLabel").grid(row=6, column=0, padx=10, pady=5, sticky="w")
default_hostname = os.popen('hostname').read().strip()
entry_hostname = ttk.Entry(root, style="TEntry")
entry_hostname.insert(0, default_hostname)
entry_hostname.grid(row=6, column=1, padx=10, pady=5, sticky="ew", ipadx=100)

# Tuổi thọ ổ cứng
ttk.Label(root, text="Tuổi thọ ổ cứng:", style="TLabel").grid(row=7, column=0, padx=10, pady=5, sticky="w")
entry_note3 = ttk.Entry(root, style="TEntry")
entry_note3.grid(row=7, column=1, padx=10, pady=5, sticky="ew", ipadx=100)
update_disk_health()

# Tuổi thọ Pin
ttk.Label(root, text="Tuổi thọ Pin:", style="TLabel").grid(row=8, column=0, padx=10, pady=5, sticky="w")
entry_pin = ttk.Entry(root, style="TEntry")
entry_pin.grid(row=8, column=1, padx=10, pady=5, sticky="ew", ipadx=100)

# Nút gửi
ttk.Button(root, text="Gửi dữ liệu", command=fast_send_file, style="TButton").grid(row=9, column=0, padx=10, pady=10, sticky="ew")
ttk.Button(root, text="Tạo", command=submit_form, style="TButton").grid(row=9, column=1, padx=10, pady=10, sticky="ew")

os.system(cmd_create_report_pin)
root.after(2000, update_battery_status)

root.columnconfigure(1, weight=1)
root.mainloop()